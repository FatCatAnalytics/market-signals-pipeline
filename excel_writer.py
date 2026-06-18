"""
excel_writer.py — pixel-perfect openpyxl output for market_signals_pipeline
============================================================================
Generates market_signals_report.xlsx with:

  Sheet 1  — Signal Summary Dashboard  (A–K)
  Sheet 2  — Sector Changes
  Sheet 3  — HQ Changes               (extra USA Region col)
  Sheet 4  — M&A & Spinoffs
  Sheet 5  — Renaming Rebranding
  Sheet 6  — Operational Changes
  Sheet 7  — Bankruptcy
  Sheet 8  — Pipeline Stats

All colours, widths, fonts, borders and freeze-pane settings are pulled from
config.py so they stay in sync with the reference file.
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

import config
from classifier import SignalResult


# ─────────────────────────────────────────────────────────────────────────────
# Shared style helpers
# ─────────────────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _font(
    color: str,
    size: int  = 11,
    bold: bool = False,
    name: str  = "Calibri",
) -> Font:
    return Font(name=name, size=size, bold=bold, color=color)


def _border() -> Border:
    s = Side(style="thin", color=config.COLOR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _align(horizontal: str = "left", wrap: bool = False) -> Alignment:
    return Alignment(
        horizontal=horizontal,
        vertical="top",
        wrap_text=wrap,
    )


# Pre-built style objects (reused across rows — openpyxl copies on assign)
_HEADER_FILL   = _fill(config.COLOR_HEADER_FILL)
_ALT_FILL      = _fill(config.COLOR_ALT_ROW)
_WHITE_FILL    = _fill(config.COLOR_WHITE_ROW)

_HEADER_FONT   = _font("FFFFFF", size=11, bold=True)
_COMPANY_FONT  = _font(config.COLOR_COMPANY_TEXT, size=11, bold=True)
_BODY_FONT     = _font(config.COLOR_BODY_TEXT,    size=11)
_SOURCES_FONT  = _font(config.COLOR_SOURCES_TEXT, size=9)
_REGION_FONT   = _font(config.COLOR_REGION_TEXT,  size=10, bold=True)
_TOTAL_FONT    = _font(config.COLOR_TOTAL_TEXT,   size=12, bold=True)
_SUMMARY_FONT  = _font(config.COLOR_BODY_TEXT,    size=10)
_TICK_FONT     = _font(config.COLOR_TICK,          size=13, bold=True)
_DASH_FONT     = _font(config.COLOR_DASH,          size=11, bold=False)

_BORDER        = _border()
_HEADER_ALIGN  = _align("center")
_LEFT_WRAP     = _align("left",   wrap=True)
_LEFT_TOP      = _align("left",   wrap=False)
_CENTER_TOP    = _align("center", wrap=False)


def _style_cell(cell, font, fill, alignment, border=None):
    cell.font      = font
    cell.fill      = fill
    cell.alignment = alignment
    if border is not None:
        cell.border = border


def _row_fill(row_idx: int) -> PatternFill:
    """Odd data rows get alt blue; even rows stay white."""
    return _ALT_FILL if row_idx % 2 == 1 else _WHITE_FILL


def _tick_dash(flag: bool) -> tuple[str, Font]:
    return ("✓", _TICK_FONT) if flag else ("—", _DASH_FONT)


def _set_col_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────────────────────────────────────────────────────────
# Dashboard sheet
# ─────────────────────────────────────────────────────────────────────────────
_DASH_HEADERS = [
    "Company",
    "Sector/\nSubsector Change",
    "HQ/Domicile\nChange",
    "M&A /\nSpinoffs",
    "Renaming /\nRebranding",
    "Operational\nChange/Shutdown",
    "Bankruptcy /\nLiquidation",
    "Total\nSignals",
    f"Signal Summary  ({config.DATE_RANGE})",
    "HQ Move\nUSA Region",
    "Confirmed\nShutdown",
]


def _build_dashboard(wb: Workbook, results: List[SignalResult]) -> None:
    ws = wb.active
    ws.title = "Signal Summary Dashboard"

    _set_col_widths(ws, config.DASH_COL_WIDTHS)
    ws.row_dimensions[1].height = config.HEADER_ROW_HEIGHT
    ws.freeze_panes = "A2"

    # ── header row ────────────────────────────────────────────────────────────
    for col, label in enumerate(_DASH_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        _style_cell(cell, _HEADER_FONT, _HEADER_FILL, _HEADER_ALIGN, _BORDER)
        cell.alignment = Alignment(
            horizontal="center", vertical="center",
            wrap_text=True,
        )

    # ── data rows ─────────────────────────────────────────────────────────────
    for r_idx, res in enumerate(results, start=1):
        row_num  = r_idx + 1          # excel row (header = 1)
        row_fill = _row_fill(r_idx)   # alternating

        ws.row_dimensions[row_num].height = config.DATA_ROW_HEIGHT

        # A — Company name
        c = ws.cell(row=row_num, column=1, value=res.company)
        _style_cell(c, _COMPANY_FONT, row_fill, _LEFT_TOP, _BORDER)

        # B–G — tick / dash columns
        flags = [
            res.sector_change,
            res.hq_change,
            res.ma_spinoff,
            res.renaming,
            res.operational_change,
            res.bankruptcy,
        ]
        for col_offset, flag in enumerate(flags, start=2):
            val, fnt = _tick_dash(flag)
            c = ws.cell(row=row_num, column=col_offset, value=val)
            _style_cell(c, fnt, row_fill, _CENTER_TOP, _BORDER)

        # H — Total Signals
        c = ws.cell(row=row_num, column=8, value=res.total_signals)
        _style_cell(c, _TOTAL_FONT, row_fill,
                    Alignment(horizontal="center", vertical="top"), _BORDER)

        # I — Signal Summary
        c = ws.cell(row=row_num, column=9, value=res.summary)
        _style_cell(c, _SUMMARY_FONT, row_fill, _LEFT_WRAP, _BORDER)

        # J — HQ USA Region
        region_val = res.hq_region if res.hq_change else "—"
        region_fnt = _REGION_FONT if res.hq_change else _DASH_FONT
        c = ws.cell(row=row_num, column=10, value=region_val)
        _style_cell(c, region_fnt, row_fill, _CENTER_TOP, _BORDER)

        # K — Confirmed Shutdown
        val, fnt = _tick_dash(res.shutdown)
        c = ws.cell(row=row_num, column=11, value=val)
        _style_cell(c, fnt, row_fill, _CENTER_TOP, _BORDER)

    # Auto-filter on all data columns
    last_row = 1 + len(results)
    ws.auto_filter.ref = f"A1:K{last_row}"


# ─────────────────────────────────────────────────────────────────────────────
# Generic detail sheet  (Company | Detail | Sources)
# ─────────────────────────────────────────────────────────────────────────────
def _build_detail_sheet(
    wb:       Workbook,
    title:    str,
    headers:  list[str],
    widths:   list[float],
    rows:     list[tuple],      # each tuple = (company, detail, sources_str) or (company, region, detail, sources_str)
    tab_color: str = "1F4E79",
) -> None:
    ws = wb.create_sheet(title=title)
    ws.sheet_properties.tabColor = tab_color

    _set_col_widths(ws, widths)
    ws.row_dimensions[1].height = config.HEADER_ROW_HEIGHT
    ws.freeze_panes = "A2"

    # header
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        _style_cell(cell, _HEADER_FONT, _HEADER_FILL,
                    Alignment(horizontal="center", vertical="center", wrap_text=True),
                    _BORDER)

    if not rows:
        ws.cell(row=2, column=1, value="No signals detected.")
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        return

    for r_idx, row_data in enumerate(rows, start=1):
        row_num  = r_idx + 1
        row_fill = _row_fill(r_idx)

        ws.row_dimensions[row_num].height = config.DATA_ROW_HEIGHT

        n_cols = len(row_data)
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            is_first = col_idx == 1
            is_last  = col_idx == n_cols

            if is_first:
                fnt   = _COMPANY_FONT
                align = _LEFT_TOP
            elif is_last:
                fnt   = _SOURCES_FONT
                align = _LEFT_WRAP
            else:
                fnt   = _BODY_FONT
                align = _LEFT_WRAP

            _style_cell(cell, fnt, row_fill, align, _BORDER)

    last_row = 1 + len(rows)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"


# ─────────────────────────────────────────────────────────────────────────────
# Tab-colour palette for detail sheets
# ─────────────────────────────────────────────────────────────────────────────
_TAB_COLORS = {
    "Sector Changes":       "2E75B6",
    "HQ Changes":           "70AD47",
    "M&A & Spinoffs":       "ED7D31",
    "Renaming Rebranding":  "FFC000",
    "Operational Changes":  "4472C4",
    "Bankruptcy":           "C00000",
    "Pipeline Stats":       "7030A0",
}


# ─────────────────────────────────────────────────────────────────────────────
# Pipeline Stats sheet
# ─────────────────────────────────────────────────────────────────────────────
def _build_stats_sheet(
    wb:           Workbook,
    results:      List[SignalResult],
    elapsed_sec:  float = 0.0,
    date_window:  str   = config.DATE_RANGE,
) -> None:
    ws = wb.create_sheet(title="Pipeline Stats")
    ws.sheet_properties.tabColor = _TAB_COLORS["Pipeline Stats"]

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 28
    ws.row_dimensions[1].height = config.HEADER_ROW_HEIGHT

    # Header
    for col, label in enumerate(["Metric", "Value"], start=1):
        cell = ws.cell(row=1, column=col, value=label)
        _style_cell(cell, _HEADER_FONT, _HEADER_FILL, _HEADER_ALIGN, _BORDER)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Stats data
    total_companies = len(results)
    with_signals    = sum(1 for r in results if r.total_signals > 0)

    signal_counts = {
        "Sector Changes":      sum(1 for r in results if r.sector_change),
        "HQ Changes":          sum(1 for r in results if r.hq_change),
        "M&A / Spinoffs":      sum(1 for r in results if r.ma_spinoff),
        "Renaming/Rebranding": sum(1 for r in results if r.renaming),
        "Operational Changes": sum(1 for r in results if r.operational_change),
        "Confirmed Shutdowns": sum(1 for r in results if r.shutdown),
        "Bankruptcies":        sum(1 for r in results if r.bankruptcy),
    }

    stats_rows = [
        ("Generated",                  datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Date Window",                date_window),
        ("Total Companies Processed",  total_companies),
        ("Companies with Signals",     with_signals),
        ("",                           ""),
        ("— Signal Breakdown —",       ""),
    ] + [(k, v) for k, v in signal_counts.items()]

    if elapsed_sec > 0:
        mins, secs = divmod(int(elapsed_sec), 60)
        stats_rows.append(("Pipeline Runtime", f"{mins}m {secs}s"))

    for r_idx, (metric, value) in enumerate(stats_rows, start=1):
        row_num  = r_idx + 1
        row_fill = _row_fill(r_idx)

        ws.row_dimensions[row_num].height = 22

        if metric.startswith("—"):
            # section divider
            c = ws.cell(row=row_num, column=1, value=metric)
            _style_cell(c, _font(config.COLOR_HEADER_FILL, 11, bold=True),
                        _fill("D9E2F3"), _LEFT_TOP, _BORDER)
            ws.cell(row=row_num, column=2).fill    = _fill("D9E2F3")
            ws.cell(row=row_num, column=2).border  = _BORDER
        elif metric == "":
            # empty spacer row
            for col in [1, 2]:
                c = ws.cell(row=row_num, column=col, value="")
                c.fill   = _WHITE_FILL
                c.border = _BORDER
        else:
            c = ws.cell(row=row_num, column=1, value=metric)
            _style_cell(c, _BODY_FONT, row_fill, _LEFT_TOP, _BORDER)
            c = ws.cell(row=row_num, column=2, value=value)
            _style_cell(c, _BODY_FONT, row_fill,
                        Alignment(horizontal="right", vertical="top"), _BORDER)


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────
def write_report(
    results:     List[SignalResult],
    output_path: str   = config.DEFAULT_OUTPUT_XLS,
    elapsed_sec: float = 0.0,
) -> str:
    """
    Build the full Excel report and save to *output_path*.
    Returns the resolved absolute path.
    """
    wb = Workbook()

    # ── Sheet 1: Dashboard ────────────────────────────────────────────────────
    _build_dashboard(wb, results)
    wb.active.sheet_properties.tabColor = "1F4E79"

    # ── Sheet 2: Sector Changes ───────────────────────────────────────────────
    sector_rows = [
        (r.company, r.sector_detail, "; ".join(r.sources))
        for r in results if r.sector_change
    ]
    _build_detail_sheet(
        wb, "Sector Changes",
        headers=["Company", f"Sector Change Detail  ({config.DATE_RANGE})", "Sources"],
        widths=config.DETAIL_COL_WIDTHS_DEFAULT,
        rows=sector_rows,
        tab_color=_TAB_COLORS["Sector Changes"],
    )

    # ── Sheet 3: HQ Changes ───────────────────────────────────────────────────
    hq_rows = [
        (r.company, r.hq_region or "—", r.hq_detail, "; ".join(r.sources))
        for r in results if r.hq_change
    ]
    _build_detail_sheet(
        wb, "HQ Changes",
        headers=["Company", "USA Region", f"HQ / Domicile Change  ({config.DATE_RANGE})", "Sources"],
        widths=config.DETAIL_COL_WIDTHS_HQ,
        rows=hq_rows,
        tab_color=_TAB_COLORS["HQ Changes"],
    )

    # ── Sheet 4: M&A & Spinoffs ───────────────────────────────────────────────
    ma_rows = [
        (r.company, r.ma_detail, "; ".join(r.sources))
        for r in results if r.ma_spinoff
    ]
    _build_detail_sheet(
        wb, "M&A & Spinoffs",
        headers=["Company", f"M&A / Spinoff Detail  ({config.DATE_RANGE})", "Sources"],
        widths=config.DETAIL_COL_WIDTHS_DEFAULT,
        rows=ma_rows,
        tab_color=_TAB_COLORS["M&A & Spinoffs"],
    )

    # ── Sheet 5: Renaming Rebranding ──────────────────────────────────────────
    rename_rows = [
        (r.company, r.rename_detail, "; ".join(r.sources))
        for r in results if r.renaming
    ]
    _build_detail_sheet(
        wb, "Renaming Rebranding",
        headers=["Company", f"Renaming / Rebranding Detail  ({config.DATE_RANGE})", "Sources"],
        widths=config.DETAIL_COL_WIDTHS_DEFAULT,
        rows=rename_rows,
        tab_color=_TAB_COLORS["Renaming Rebranding"],
    )

    # ── Sheet 6: Operational Changes ─────────────────────────────────────────
    ops_rows = [
        (r.company, r.ops_detail, "; ".join(r.sources))
        for r in results if r.operational_change or r.shutdown
    ]
    _build_detail_sheet(
        wb, "Operational Changes",
        headers=["Company", f"Operational Change / Shutdown Detail  ({config.DATE_RANGE})", "Sources"],
        widths=config.DETAIL_COL_WIDTHS_DEFAULT,
        rows=ops_rows,
        tab_color=_TAB_COLORS["Operational Changes"],
    )

    # ── Sheet 7: Bankruptcy ───────────────────────────────────────────────────
    bk_rows = [
        (r.company, r.bankruptcy_detail, "; ".join(r.sources))
        for r in results if r.bankruptcy
    ]
    _build_detail_sheet(
        wb, "Bankruptcy",
        headers=["Company", f"Bankruptcy / Liquidation Detail  ({config.DATE_RANGE})", "Sources"],
        widths=config.DETAIL_COL_WIDTHS_DEFAULT,
        rows=bk_rows,
        tab_color=_TAB_COLORS["Bankruptcy"],
    )

    # ── Sheet 8: Pipeline Stats ───────────────────────────────────────────────
    _build_stats_sheet(wb, results, elapsed_sec=elapsed_sec)

    # ── Save ──────────────────────────────────────────────────────────────────
    abs_path = os.path.abspath(output_path)
    os.makedirs(os.path.dirname(abs_path) if os.path.dirname(abs_path) else ".", exist_ok=True)
    wb.save(abs_path)
    return abs_path
